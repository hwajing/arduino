import pandas as pd
import os
from openpyxl import load_workbook
# ===============================================================================================================
# 技术文档
# 【pandas】
# https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.merge.html?highlight=merge#pandas.merge
# 【openpyxl】
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html
# ===============================================================================================================
# ★★★请自行修改相应参数★★★
# 【filePath】文件路径，不包含文件名
# 【fileName】文件名（含扩展名）
# 【sheetLeft】【sheetRight】2个sheet名，以Left表为基准，将Right表的数据匹配过来。
# 【joinColumns】匹配列，左右表必须同名
# 【resultColumns】为Right表中挑选列（可指定多列，但必须包含匹配列）。显示结果为左表所有列+右表挑选列。
# 【fileName4Result】如需在新文件中显示结果，此参数指定新文件名（如与原文件名一致则在原文件内添加sheet）
filePath = 'C:/Users/hwaji/PycharmProjects/Arduino/excel/'
fileName = 'vlookup.xlsx'
sheetLeft = 'sheet1'
sheetRight = 'SAP主数据'
joinColumns = ['客户代码'] # 匹配列为复数列时['col1', 'col2']
resultColumns = ['客户代码', 'SAP客户代码']
fileName4Result = 'result.xlsx'
sheetName4Result = 'result'
# ★★★请自行修改相应参数★★★
# ==============================================================================================
# 左连匹配
how = 'left'
fileName = filePath + fileName
fileName4Result = filePath + fileName4Result
# ==============================================================================================

if not os.path.exists(fileName):
    print('File not exist: ' + fileName)
else:
    try:
        # 左表
        dataLeft = pd.read_excel(fileName, encoding='utf-8', sheet_name=sheetLeft)
        print(dataLeft)
        # 右表
        dataRight = pd.read_excel(fileName, encoding='utf-8', sheet_name=sheetRight)
        print(dataRight)
        # 结合后（右表仅包含挑选字段）
        all_data = dataLeft.merge(dataRight[resultColumns], on=joinColumns, how=how)
        print(all_data)

        # ==========================================================================
        # 删除重名文件
        if os.path.exists(fileName4Result) and fileName4Result != fileName:
            os.remove(fileName4Result)

        writer = pd.ExcelWriter(fileName4Result, engine='openpyxl')
        # 如果在原文件中写
        if fileName4Result == fileName:
            book = load_workbook(fileName4Result)
            # 删除结果sheet
            if sheetName4Result in book.sheetnames:
                std = book[sheetName4Result]
                book.remove(std)
            writer.book = book

        # 写文件
        all_data.to_excel(writer, sheet_name=sheetName4Result)
        writer.save()

    except Exception as ex:
        print(ex.message)
